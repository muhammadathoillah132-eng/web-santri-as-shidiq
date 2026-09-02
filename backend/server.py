"""AS SHIDIQ SANTRI MANAGEMENT — backend."""
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, io, uuid, logging, calendar
import bcrypt
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Any
from datetime import datetime, timezone, timedelta
from reportlab.lib.pagesizes import A5
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
OWNER_EMAIL = os.environ.get('OWNER_EMAIL', 'akayfikanita@gmail.com').lower()

app = FastAPI()
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# =============== Helpers ===============
def now_utc():
    return datetime.now(timezone.utc)


def iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def new_id(prefix="id"):
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


async def log_activity(user, action: str, entity: str, detail: str = ""):
    await db.activity_logs.insert_one({
        "log_id": new_id("log"),
        "user_id": user.get("user_id"),
        "user_email": user.get("email"),
        "user_name": user.get("name"),
        "action": action,
        "entity": entity,
        "detail": detail,
        "created_at": iso(now_utc()),
    })


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("session_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    exp = session["expires_at"]
    if isinstance(exp, str):
        exp = datetime.fromisoformat(exp)
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp < now_utc():
        raise HTTPException(status_code=401, detail="Session expired")
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    if user.get("status") == "inactive":
        raise HTTPException(status_code=403, detail="Account disabled")
    return user


def require_super_admin(user: dict) -> dict:
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Requires super_admin")
    return user


# =============== AUTH ===============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


class LoginIn(BaseModel):
    identifier: str  # username atau email
    password: str
    remember_me: bool = False


@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    ident = payload.identifier.strip().lower()
    user = await db.users.find_one({"$or": [{"email": ident}, {"username": ident}]}, {"_id": 0})
    invalid = HTTPException(401, "Username/email atau password salah")
    if not user or not user.get("password_hash"):
        raise invalid
    la = await db.login_attempts.find_one({"identifier": ident})
    if la and la.get("count", 0) >= 5 and la.get("locked_until", "") > iso(now_utc()):
        raise HTTPException(429, "Terlalu banyak percobaan login. Coba lagi 15 menit lagi.")
    if not verify_password(payload.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1},
             "$set": {"locked_until": iso(now_utc() + timedelta(minutes=15)), "last_attempt": iso(now_utc())}},
            upsert=True,
        )
        raise invalid
    if user.get("status") == "inactive":
        raise HTTPException(403, "Akun dinonaktifkan. Hubungi Super Admin.")
    await db.login_attempts.delete_one({"identifier": ident})
    days = 30 if payload.remember_me else 1
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "user_id": user["user_id"], "session_token": session_token,
        "expires_at": iso(now_utc() + timedelta(days=days)),
        "created_at": iso(now_utc()),
    })
    response.set_cookie(
        key="session_token", value=session_token,
        max_age=days * 24 * 3600, httponly=True, secure=True, samesite="none", path="/",
    )
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"last_login": iso(now_utc())}})
    await log_activity(user, "login", "auth", f"{user.get('email')} login")
    user.pop("password_hash", None)
    return {"user": user, "session_token": session_token}


@api.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    return user


@api.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get("session_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/", secure=True, samesite="none")
    return {"ok": True}


# =============== ADMINS ===============
@api.get("/admins")
async def list_admins(user=Depends(get_current_user)):
    require_super_admin(user)
    return await db.users.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


class AdminIn(BaseModel):
    name: str
    email: str
    username: Optional[str] = None
    password: Optional[str] = None
    whatsapp: Optional[str] = ""
    role: str = "admin"
    status: str = "active"


@api.post("/admins")
async def add_admin(payload: AdminIn, user=Depends(get_current_user)):
    require_super_admin(user)
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already exists")
    if not payload.password or len(payload.password) < 6:
        raise HTTPException(400, "Password wajib diisi (minimal 6 karakter)")
    username = (payload.username or email.split("@")[0]).strip().lower()
    if await db.users.find_one({"username": username}):
        raise HTTPException(400, "Username sudah dipakai")
    uid = new_id("user")
    doc = {
        "user_id": uid, "email": email, "name": payload.name,
        "username": username, "password_hash": hash_password(payload.password),
        "whatsapp": payload.whatsapp, "role": payload.role,
        "status": payload.status,
        "picture": "", "created_at": iso(now_utc()), "last_login": None,
    }
    await db.users.insert_one(doc)
    await log_activity(user, "create_admin", "admin", f"Added admin {email}")
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    return doc


@api.post("/admins/{user_id}/reset-password")
async def reset_admin_password(user_id: str, payload: dict, user=Depends(get_current_user)):
    require_super_admin(user)
    new_pw = (payload.get("password") or "").strip()
    if len(new_pw) < 6:
        raise HTTPException(400, "Password minimal 6 karakter")
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Not found")
    await db.users.update_one({"user_id": user_id}, {"$set": {"password_hash": hash_password(new_pw)}})
    await db.user_sessions.delete_many({"user_id": user_id})
    await log_activity(user, "reset_password", "admin", f"Reset password {target.get('email')}")
    return {"ok": True}


@api.patch("/admins/{user_id}")
async def update_admin(user_id: str, payload: dict, user=Depends(get_current_user)):
    require_super_admin(user)
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Not found")
    if target.get("email") == OWNER_EMAIL and payload.get("role") and payload["role"] != "super_admin":
        raise HTTPException(400, "Cannot change owner role")
    allowed = {k: v for k, v in payload.items() if k in {"name", "whatsapp", "role", "status"}}
    await db.users.update_one({"user_id": user_id}, {"$set": allowed})
    await log_activity(user, "update_admin", "admin", f"Updated {target.get('email')}")
    return await db.users.find_one({"user_id": user_id}, {"_id": 0})


@api.delete("/admins/{user_id}")
async def delete_admin(user_id: str, user=Depends(get_current_user)):
    require_super_admin(user)
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Not found")
    if target.get("email") == OWNER_EMAIL:
        raise HTTPException(400, "Cannot delete owner")
    await db.users.delete_one({"user_id": user_id})
    await db.user_sessions.delete_many({"user_id": user_id})
    await log_activity(user, "delete_admin", "admin", f"Deleted {target.get('email')}")
    return {"ok": True}


# =============== MASTER DATA ===============
MASTER_TYPES = {"programs", "kelas", "jurusan", "tahun_ajaran", "payment_types", "asramas", "kamars"}


@api.get("/master/{mtype}")
async def list_master(mtype: str, user=Depends(get_current_user)):
    if mtype not in MASTER_TYPES:
        raise HTTPException(404, "Unknown master type")
    return await db[f"m_{mtype}"].find({}, {"_id": 0}).sort("name", 1).to_list(1000)


@api.post("/master/{mtype}")
async def add_master(mtype: str, payload: dict, user=Depends(get_current_user)):
    if mtype not in MASTER_TYPES:
        raise HTTPException(404, "Unknown master type")
    doc = {**payload, "id": new_id("m"), "created_at": iso(now_utc())}
    await db[f"m_{mtype}"].insert_one(doc)
    doc.pop("_id", None)
    await log_activity(user, "create_master", mtype, doc.get("name", ""))
    return doc


@api.patch("/master/{mtype}/{item_id}")
async def upd_master(mtype: str, item_id: str, payload: dict, user=Depends(get_current_user)):
    if mtype not in MASTER_TYPES:
        raise HTTPException(404, "Unknown master type")
    await db[f"m_{mtype}"].update_one({"id": item_id}, {"$set": payload})
    await log_activity(user, "update_master", mtype, item_id)
    return await db[f"m_{mtype}"].find_one({"id": item_id}, {"_id": 0})


@api.delete("/master/{mtype}/{item_id}")
async def del_master(mtype: str, item_id: str, user=Depends(get_current_user)):
    if mtype not in MASTER_TYPES:
        raise HTTPException(404, "Unknown master type")
    require_super_admin(user)
    await db[f"m_{mtype}"].delete_one({"id": item_id})
    await log_activity(user, "delete_master", mtype, item_id)
    return {"ok": True}


# =============== SANTRI ===============
@api.get("/santri")
async def list_santri(
    q: Optional[str] = None, gender: Optional[str] = None,
    program: Optional[str] = None, kelas: Optional[str] = None,
    status: Optional[str] = None, asrama: Optional[str] = None,
    user=Depends(get_current_user),
):
    filt = {}
    if q:
        rx = {"$regex": q, "$options": "i"}
        filt["$or"] = [{"nama": rx}, {"nomor_induk": rx}, {"nik": rx}, {"whatsapp": rx}, {"nisn": rx}]
    if gender: filt["gender"] = gender
    if program: filt["program"] = program
    if kelas: filt["kelas"] = kelas
    if status: filt["status"] = status
    if asrama: filt["asrama"] = asrama
    return await db.santri.find(filt, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.get("/santri/{sid}")
async def get_santri(sid: str, user=Depends(get_current_user)):
    s = await db.santri.find_one({"santri_id": sid}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Santri not found")
    payments = await db.payments.find({"santri_id": sid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    invoices = await db.invoices.find({"santri_id": sid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"santri": s, "payments": payments, "invoices": invoices}


@api.post("/santri")
async def create_santri(payload: dict, user=Depends(get_current_user)):
    if payload.get("nomor_induk"):
        exist = await db.santri.find_one({"nomor_induk": payload["nomor_induk"]})
        if exist:
            raise HTTPException(400, "Nomor Induk sudah terdaftar")
    doc = {
        **payload,
        "santri_id": new_id("s"),
        "status": payload.get("status", "aktif"),
        "created_at": iso(now_utc()),
        "updated_at": iso(now_utc()),
    }
    await db.santri.insert_one(doc)
    doc.pop("_id", None)
    await log_activity(user, "create_santri", "santri", f"{doc.get('nama')} ({doc.get('nomor_induk')})")
    return doc


@api.patch("/santri/{sid}")
async def upd_santri(sid: str, payload: dict, user=Depends(get_current_user)):
    payload["updated_at"] = iso(now_utc())
    await db.santri.update_one({"santri_id": sid}, {"$set": payload})
    await log_activity(user, "update_santri", "santri", sid)
    return await db.santri.find_one({"santri_id": sid}, {"_id": 0})


@api.delete("/santri/{sid}")
async def del_santri(sid: str, user=Depends(get_current_user)):
    require_super_admin(user)
    s = await db.santri.find_one({"santri_id": sid}, {"_id": 0})
    if not s: raise HTTPException(404, "Not found")
    await db.santri.delete_one({"santri_id": sid})
    await log_activity(user, "delete_santri", "santri", f"{s.get('nama')}")
    return {"ok": True}


# =============== INVOICES (TAGIHAN) ===============
@api.get("/tagihan")
async def list_tagihan(status: Optional[str] = None, santri_id: Optional[str] = None,
                        user=Depends(get_current_user)):
    filt = {}
    if status: filt["status"] = status
    if santri_id: filt["santri_id"] = santri_id
    return await db.invoices.find(filt, {"_id": 0}).sort("created_at", -1).to_list(5000)


class InvoiceBatchIn(BaseModel):
    jenis: str  # payment type name
    periode: str  # e.g. "September 2026"
    nominal: float
    target_type: str  # "all" | "kelas" | "program" | "santri"
    target_ids: List[str] = []
    keterangan: Optional[str] = ""


@api.post("/tagihan")
async def create_tagihan(payload: InvoiceBatchIn, user=Depends(get_current_user)):
    filt = {"status": "aktif"}
    if payload.target_type == "kelas":
        filt["kelas"] = {"$in": payload.target_ids}
    elif payload.target_type == "program":
        filt["program"] = {"$in": payload.target_ids}
    elif payload.target_type == "santri":
        filt = {"santri_id": {"$in": payload.target_ids}}
    santris = await db.santri.find(filt, {"_id": 0, "santri_id": 1, "nama": 1, "nomor_induk": 1}).to_list(5000)
    docs = []
    for s in santris:
        docs.append({
            "invoice_id": new_id("inv"),
            "santri_id": s["santri_id"],
            "santri_nama": s.get("nama"),
            "santri_nomor_induk": s.get("nomor_induk"),
            "jenis": payload.jenis,
            "periode": payload.periode,
            "nominal": payload.nominal,
            "terbayar": 0,
            "status": "belum_lunas",
            "keterangan": payload.keterangan,
            "created_at": iso(now_utc()),
            "created_by": user.get("email"),
        })
    if docs:
        await db.invoices.insert_many(docs)
    await log_activity(user, "create_tagihan", "invoice",
                        f"{payload.jenis} {payload.periode} untuk {len(docs)} santri")
    return {"created": len(docs)}


@api.patch("/tagihan/{iid}")
async def upd_tagihan(iid: str, payload: dict, user=Depends(get_current_user)):
    await db.invoices.update_one({"invoice_id": iid}, {"$set": payload})
    await log_activity(user, "update_tagihan", "invoice", iid)
    return await db.invoices.find_one({"invoice_id": iid}, {"_id": 0})


@api.delete("/tagihan/{iid}")
async def del_tagihan(iid: str, user=Depends(get_current_user)):
    require_super_admin(user)
    await db.invoices.delete_one({"invoice_id": iid})
    await log_activity(user, "delete_tagihan", "invoice", iid)
    return {"ok": True}


# =============== PAYMENTS ===============
class PaymentIn(BaseModel):
    santri_id: str
    invoice_id: Optional[str] = None
    jenis: str
    periode: Optional[str] = ""
    nominal: float
    metode: str = "cash"
    keterangan: Optional[str] = ""
    tanggal: Optional[str] = None


@api.get("/pembayaran")
async def list_payments(
    q: Optional[str] = None, jenis: Optional[str] = None,
    status: Optional[str] = None, month: Optional[str] = None,
    user=Depends(get_current_user),
):
    filt = {}
    if jenis: filt["jenis"] = jenis
    if status: filt["status"] = status
    if q:
        rx = {"$regex": q, "$options": "i"}
        filt["$or"] = [{"santri_nama": rx}, {"santri_nomor_induk": rx}, {"trx_no": rx}]
    if month:
        filt["periode_bulan"] = month
    return await db.payments.find(filt, {"_id": 0}).sort("created_at", -1).to_list(5000)


@api.post("/pembayaran")
async def create_payment(payload: PaymentIn, user=Depends(get_current_user)):
    s = await db.santri.find_one({"santri_id": payload.santri_id}, {"_id": 0})
    if not s: raise HTTPException(404, "Santri not found")
    tgl = payload.tanggal or iso(now_utc())
    trx_no = f"TRX-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    doc = {
        "payment_id": new_id("pay"),
        "trx_no": trx_no,
        "santri_id": payload.santri_id,
        "santri_nama": s.get("nama"),
        "santri_nomor_induk": s.get("nomor_induk"),
        "invoice_id": payload.invoice_id,
        "jenis": payload.jenis,
        "periode": payload.periode,
        "periode_bulan": (payload.periode or "")[:20],
        "nominal": payload.nominal,
        "metode": payload.metode,
        "keterangan": payload.keterangan,
        "tanggal": tgl,
        "status": "lunas",
        "admin_email": user.get("email"),
        "admin_name": user.get("name"),
        "created_at": iso(now_utc()),
    }
    await db.payments.insert_one(doc)
    # Update invoice if tied
    if payload.invoice_id:
        inv = await db.invoices.find_one({"invoice_id": payload.invoice_id}, {"_id": 0})
        if inv:
            terbayar = (inv.get("terbayar", 0) or 0) + payload.nominal
            status = "lunas" if terbayar >= inv["nominal"] else "cicilan"
            await db.invoices.update_one({"invoice_id": payload.invoice_id},
                                         {"$set": {"terbayar": terbayar, "status": status}})
    doc.pop("_id", None)
    await log_activity(user, "create_payment", "payment",
                        f"{trx_no} - {s.get('nama')} - Rp{payload.nominal:,.0f}")
    return doc


@api.delete("/pembayaran/{pid}")
async def del_payment(pid: str, user=Depends(get_current_user)):
    require_super_admin(user)
    p = await db.payments.find_one({"payment_id": pid}, {"_id": 0})
    if not p: raise HTTPException(404, "Not found")
    await db.payments.delete_one({"payment_id": pid})
    if p.get("invoice_id"):
        inv = await db.invoices.find_one({"invoice_id": p["invoice_id"]}, {"_id": 0})
        if inv:
            terbayar = max(0, (inv.get("terbayar", 0) or 0) - p["nominal"])
            status = "belum_lunas" if terbayar == 0 else ("lunas" if terbayar >= inv["nominal"] else "cicilan")
            await db.invoices.update_one({"invoice_id": p["invoice_id"]},
                                          {"$set": {"terbayar": terbayar, "status": status}})
    await log_activity(user, "delete_payment", "payment", pid)
    return {"ok": True}


# =============== KWITANSI PDF ===============
@api.get("/pembayaran/{pid}/kwitansi")
async def kwitansi_pdf(pid: str, user=Depends(get_current_user)):
    p = await db.payments.find_one({"payment_id": pid}, {"_id": 0})
    if not p: raise HTTPException(404, "Not found")
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A5)
    W, H = A5
    emerald = colors.HexColor("#047857")
    gold = colors.HexColor("#D97706")
    slate = colors.HexColor("#1E293B")

    # Header band
    c.setFillColor(emerald)
    c.rect(0, H - 2.5 * cm, W, 2.5 * cm, fill=1, stroke=0)
    c.setFillColor(gold)
    c.rect(0, H - 2.7 * cm, W, 0.2 * cm, fill=1, stroke=0)
    logo_path = Path(__file__).parent / "assets" / "logo.png"
    if logo_path.exists():
        c.drawImage(str(logo_path), 1 * cm, H - 2.35 * cm, width=1.9 * cm, height=1.9 * cm, mask="auto")
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(3.1 * cm, H - 1.3 * cm, "PONDOK PESANTREN AS SHIDIQ")
    c.setFont("Helvetica", 9)
    c.drawString(3.1 * cm, H - 1.8 * cm, "Sistem Manajemen Santri — Kwitansi Pembayaran")
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(gold)
    c.drawRightString(W - 1 * cm, H - 1.3 * cm, "KWITANSI RESMI")

    # Body
    y = H - 3.6 * cm
    c.setFillColor(slate)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1 * cm, y, f"No. Transaksi: {p['trx_no']}")
    c.setFont("Helvetica", 9)
    c.drawRightString(W - 1 * cm, y, f"Tanggal: {(p.get('tanggal') or '')[:10]}")

    def row(label, value, yy):
        c.setFont("Helvetica", 9); c.setFillColor(slate)
        c.drawString(1 * cm, yy, label)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(4.5 * cm, yy, f": {value}")

    y -= 0.8 * cm
    row("Nama Santri", p.get("santri_nama", ""), y); y -= 0.55 * cm
    row("Nomor Induk", p.get("santri_nomor_induk", "-"), y); y -= 0.55 * cm
    row("Jenis Bayar", p.get("jenis", ""), y); y -= 0.55 * cm
    row("Periode", p.get("periode", "-"), y); y -= 0.55 * cm
    row("Metode", p.get("metode", "cash").upper(), y); y -= 0.55 * cm
    row("Keterangan", (p.get("keterangan") or "-")[:40], y); y -= 0.9 * cm

    # Nominal box
    c.setFillColor(colors.HexColor("#ECFDF5"))
    c.rect(1 * cm, y - 0.2 * cm, W - 2 * cm, 1.3 * cm, fill=1, stroke=0)
    c.setStrokeColor(emerald); c.setLineWidth(1)
    c.rect(1 * cm, y - 0.2 * cm, W - 2 * cm, 1.3 * cm, fill=0, stroke=1)
    c.setFillColor(emerald); c.setFont("Helvetica-Bold", 12)
    c.drawString(1.3 * cm, y + 0.55 * cm, "TOTAL DIBAYAR")
    c.setFont("Helvetica-Bold", 16); c.setFillColor(gold)
    c.drawRightString(W - 1.3 * cm, y + 0.4 * cm, f"Rp {p['nominal']:,.0f}")

    # Footer / signature
    c.setFillColor(slate); c.setFont("Helvetica", 8)
    c.drawString(1 * cm, 2.5 * cm, "Dicatat oleh Admin:")
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1 * cm, 2.1 * cm, p.get("admin_name") or p.get("admin_email") or "-")
    c.setFont("Helvetica", 8); c.setFillColor(colors.grey)
    c.drawString(1 * cm, 1.5 * cm, "Terima kasih atas pembayarannya. Semoga Allah SWT")
    c.drawString(1 * cm, 1.2 * cm, "memudahkan langkah putra/putri kita menuntut ilmu.")
    c.setFillColor(emerald); c.setFont("Helvetica-Oblique", 8)
    c.drawRightString(W - 1 * cm, 1.2 * cm, "AS SHIDIQ SANTRI MANAGEMENT")
    c.showPage(); c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"inline; filename=kwitansi-{p['trx_no']}.pdf"})


# =============== DASHBOARD / REPORTS ===============
@api.get("/dashboard/stats")
async def dashboard_stats(user=Depends(get_current_user)):
    now = now_utc()
    month_prefix = now.strftime("%Y-%m")
    total = await db.santri.count_documents({})
    putra = await db.santri.count_documents({"gender": "L"})
    putri = await db.santri.count_documents({"gender": "P"})
    aktif = await db.santri.count_documents({"status": "aktif"})
    nonaktif = await db.santri.count_documents({"status": "nonaktif"})
    baru_start = iso((now - timedelta(days=30)))
    baru = await db.santri.count_documents({"created_at": {"$gte": baru_start}})
    # Payments this month
    pipe = [
        {"$match": {"tanggal": {"$regex": f"^{month_prefix}"}}},
        {"$group": {"_id": None, "total": {"$sum": "$nominal"}}}
    ]
    agg = await db.payments.aggregate(pipe).to_list(1)
    bulan_ini = agg[0]["total"] if agg else 0
    belum = await db.invoices.count_documents({"status": {"$in": ["belum_lunas", "cicilan"]}})
    return {
        "total_santri": total, "santri_putra": putra, "santri_putri": putri,
        "santri_aktif": aktif, "santri_nonaktif": nonaktif,
        "pembayaran_bulan_ini": bulan_ini, "tagihan_belum_lunas": belum,
        "santri_baru": baru,
    }


@api.get("/dashboard/charts")
async def dashboard_charts(user=Depends(get_current_user)):
    # By program
    prog = await db.santri.aggregate([
        {"$group": {"_id": "$program", "count": {"$sum": 1}}}
    ]).to_list(50)
    # By gender
    gen = await db.santri.aggregate([
        {"$group": {"_id": "$gender", "count": {"$sum": 1}}}
    ]).to_list(10)
    # Payments per month (last 6 months)
    months = []
    for i in range(5, -1, -1):
        d = (now_utc().replace(day=1) - timedelta(days=30 * i))
        m = d.strftime("%Y-%m")
        agg = await db.payments.aggregate([
            {"$match": {"tanggal": {"$regex": f"^{m}"}}},
            {"$group": {"_id": None, "total": {"$sum": "$nominal"}}}
        ]).to_list(1)
        months.append({"month": d.strftime("%b %Y"), "total": agg[0]["total"] if agg else 0})
    return {
        "by_program": [{"name": p["_id"] or "Lainnya", "count": p["count"]} for p in prog],
        "by_gender": [{"name": "Putra" if g["_id"] == "L" else ("Putri" if g["_id"] == "P" else "Lainnya"),
                        "count": g["count"]} for g in gen],
        "payments_monthly": months,
    }


@api.get("/laporan")
async def laporan(start: Optional[str] = None, end: Optional[str] = None,
                   user=Depends(get_current_user)):
    filt = {}
    if start or end:
        rng = {}
        if start: rng["$gte"] = start
        if end: rng["$lte"] = end + "T23:59:59Z"
        filt["tanggal"] = rng
    pays = await db.payments.find(filt, {"_id": 0}).sort("tanggal", -1).to_list(10000)
    total = sum(p["nominal"] for p in pays)
    # By jenis
    by_jenis = {}
    for p in pays:
        by_jenis[p["jenis"]] = by_jenis.get(p["jenis"], 0) + p["nominal"]
    return {
        "total": total, "count": len(pays),
        "by_jenis": [{"name": k, "total": v} for k, v in by_jenis.items()],
        "items": pays[:500],
    }


# =============== SEARCH ===============
@api.get("/search")
async def global_search(q: str, user=Depends(get_current_user)):
    if not q or len(q) < 2: return {"santri": []}
    rx = {"$regex": q, "$options": "i"}
    santris = await db.santri.find(
        {"$or": [{"nama": rx}, {"nomor_induk": rx}, {"nik": rx}, {"whatsapp": rx}, {"nisn": rx}]},
        {"_id": 0, "santri_id": 1, "nama": 1, "nomor_induk": 1, "kelas": 1, "program": 1, "foto": 1}
    ).limit(10).to_list(10)
    return {"santri": santris}


# =============== ACTIVITY LOGS ===============
@api.get("/aktivitas")
async def list_activity(user=Depends(get_current_user)):
    require_super_admin(user)
    return await db.activity_logs.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)


# =============== EXPORT / IMPORT ===============
@api.get("/santri/export/xlsx")
async def export_xlsx(user=Depends(get_current_user)):
    santris = await db.santri.find({}, {"_id": 0}).to_list(10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Santri"
    cols = ["nomor_induk", "nik", "nisn", "nama", "gender", "tempat_lahir", "tanggal_lahir",
            "alamat", "desa", "kecamatan", "kabupaten", "provinsi", "whatsapp", "email",
            "program", "kelas", "jurusan", "tahun_masuk", "asrama", "kamar",
            "nama_ayah", "nama_ibu", "nama_wali", "wa_wali", "status"]
    ws.append([c.upper() for c in cols])
    for s in santris:
        ws.append([s.get(c, "") for c in cols])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=santri.xlsx"},
    )


@api.get("/santri/template/xlsx")
async def template_xlsx(user=Depends(get_current_user)):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Template"
    cols = ["nomor_induk", "nik", "nisn", "nama", "gender (L/P)", "tempat_lahir", "tanggal_lahir",
            "alamat", "desa", "kecamatan", "kabupaten", "provinsi", "whatsapp", "email",
            "program", "kelas", "jurusan", "tahun_masuk", "asrama", "kamar",
            "nama_ayah", "nama_ibu", "nama_wali", "wa_wali", "status (aktif/nonaktif)"]
    ws.append(cols)
    ws.append(["24001", "3201xxxx", "0012345", "Nama Contoh", "L", "Jakarta", "2008-05-10",
                "Jl. Contoh", "Desa A", "Kec A", "Kab A", "Jabar", "0812xxx", "-",
                "SMPI", "7A", "-", "2024", "Al-Falah", "A1",
                "Ayah Contoh", "Ibu Contoh", "-", "-", "aktif"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-santri.xlsx"},
    )


@api.post("/santri/import")
async def import_xlsx(file: UploadFile = File(...), user=Depends(get_current_user)):
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"imported": 0, "skipped": 0}
    header = [str(h).lower().split(" ")[0] if h else "" for h in rows[0]]
    imported, skipped = 0, 0
    for r in rows[1:]:
        d = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
        ni = str(d.get("nomor_induk") or "").strip()
        if not ni or not d.get("nama"):
            skipped += 1; continue
        if await db.santri.find_one({"nomor_induk": ni}):
            skipped += 1; continue
        doc = {k: (str(v) if v is not None else "") for k, v in d.items()}
        doc["nomor_induk"] = ni
        doc["santri_id"] = new_id("s")
        doc["status"] = doc.get("status") or "aktif"
        doc["created_at"] = iso(now_utc())
        doc["updated_at"] = iso(now_utc())
        await db.santri.insert_one(doc)
        imported += 1
    await log_activity(user, "import_santri", "santri", f"Imported {imported}, skipped {skipped}")
    return {"imported": imported, "skipped": skipped}


# =============== SEED ===============
@api.post("/seed/dummy")
async def seed_dummy(user=Depends(get_current_user)):
    require_super_admin(user)
    from seed_data import run_seed
    result = await run_seed(db)
    await log_activity(user, "seed", "system", "Seeded dummy data")
    return result


@api.get("/")
async def root(): return {"app": "AS SHIDIQ SANTRI MANAGEMENT", "ok": True}


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[o for o in os.environ.get('CORS_ORIGINS', '').split(',') if o and o != '*'],
    allow_origin_regex=r"https://.*\.github\.io",
    allow_methods=["*"], allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client(): client.close()


@app.on_event("startup")
async def startup():
    # Ensure indexes
    await db.santri.create_index("santri_id", unique=True)
    await db.santri.create_index("nomor_induk")
    await db.users.create_index("email", unique=True)
    await db.user_sessions.create_index("session_token", unique=True)
    await db.login_attempts.create_index("identifier")
    # Bootstrap master data + demo if empty
    from seed_data import bootstrap
    await bootstrap(db, OWNER_EMAIL)
